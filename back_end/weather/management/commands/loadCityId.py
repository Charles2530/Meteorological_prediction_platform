from django.core.management.base import BaseCommand
import csv
from weather.models import City2CityId, Pro2City, ProGeography
from openpyxl import load_workbook

class Command(BaseCommand):
    help = 'Store catastrophic forecast data into the database'

    def add_arguments(self, parser):
        # Named (optional) arguments
        parser.add_argument(
            "--U",
            action="store_true",
            help="Refresh All CityInfo even existed already",
        )

    def handle(self, *args, **kwargs):
        try:
            f = open(r'/root/Meteorological_prediction_platform/back_end/China-City-List-latest.csv', 'r', newline='', encoding='utf-8')
        except FileNotFoundError:
            try:
                f = open(r'D:\Programing\SoftwareEngineering\Meteorological_prediction_platform\back_end\China-City-List-latest.csv', 'r', newline='', encoding='utf-8')
            except FileNotFoundError:
                pass
        reader = csv.reader(f)
        cnt = 0
        capital = dict()
        for i, row in enumerate(reader):
            if i < 2:
                continue
            cityId = row[0]
            cityName = row[9]
            if row[9].startswith(row[2]):
                capital.setdefault(row[7], row[0])
            try:
                Cityinfo = City2CityId.objects.get(cityId=cityId)
                if not kwargs["U"]:
                    continue
            except:
                Cityinfo = City2CityId(cityId=cityId, cityName=cityName)
            # try:
            Cityinfo.cityName = cityName
            Cityinfo.location = "{:.2f},{:.2f}".format(float(row[11]), float(row[12]))
            Cityinfo.save()

        print("City amount", len(City2CityId.objects.all()))
        print("Capital amount", len(capital))
        for pro, city in capital.items():
            Proinfo = Pro2City(proName=pro, cityId=city)
            Proinfo.save()

        self.stdout.write(self.style.SUCCESS(
            'City2CityId and Pro2City data stored successfully'))


        workbook = load_workbook('/root/geography.xlsx')
        sheet = workbook.active

        # print(ProGeography.objects.filter())
        for row in sheet.iter_rows(values_only=True):
            geo = ProGeography(proName=row[0], geographyInfo=row[1])
            geo.save()
            # print(ProGeography.objects.get(proName=row[0]).proName)
            # print(row)
            # print(row[0], row[1])


